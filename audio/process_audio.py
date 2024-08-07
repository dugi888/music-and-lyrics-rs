import numpy as np
import pandas as pd
import random
from openpyxl.reader.excel import load_workbook
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor, ExtraTreesRegressor, \
    RandomForestClassifier
from sklearn.inspection import permutation_importance
from sklearn.linear_model import LinearRegression, SGDRegressor, LogisticRegression, RidgeClassifier, SGDClassifier, \
    Ridge
from sklearn.metrics import accuracy_score, root_mean_squared_error, mean_absolute_error, mean_squared_error, r2_score, \
    confusion_matrix, precision_score, recall_score, f1_score, matthews_corrcoef
from sklearn.model_selection import KFold, train_test_split, GridSearchCV
from sklearn.neighbors import KNeighborsClassifier
from sklearn.preprocessing import LabelEncoder
from sklearn.svm import SVR, SVC
from sklearn.tree import DecisionTreeRegressor, DecisionTreeClassifier
from lime.lime_tabular import LimeTabularExplainer
import utils
import sklearn
import IPython
import matplotlib.pyplot as plt
import seaborn as sns


class AudioProcessor:

    @staticmethod
    def append_df_to_excel(filename, df, sheet_name):
        try:
            # Load the existing workbook
            book = load_workbook(filename)
            if sheet_name in book.sheetnames:
                # Load existing sheet into a DataFrame
                existing_df = pd.read_excel(filename, sheet_name=sheet_name)
                # Append the new data without column names
                combined_df = pd.concat([existing_df, df], ignore_index=True)
            else:
                # If the sheet does not exist, just use the new DataFrame
                combined_df = df

            # Write to Excel
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
                if sheet_name in writer.book.sheetnames:
                    writer.book.remove(writer.book[sheet_name])
                combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
        except FileNotFoundError:
            # If the file doesn't exist, create a new one and write the DataFrame
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    from sklearn.preprocessing import LabelEncoder

    def preprocess_features(self, X):
        X_processed = X.copy()

        # Encode categorical features
        for column in X_processed.select_dtypes(include=['object']).columns:
            le = LabelEncoder()
            X_processed[column] = le.fit_transform(X_processed[column])

        return X_processed

    def preprocess_target(self, y):
        if isinstance(y, pd.Series):
            y = y.values  # Convert to ndarray

        # Check if y contains non-numeric values
        if not np.issubdtype(y.dtype, np.number):
            le = LabelEncoder()
            y_processed = le.fit_transform(y)
        else:
            y_processed = y

        return y_processed

    def evaluation_regressor(self, features_df, target_df, column_to_predict, model, param_grid, test_size=0.2):

        model_name = type(model).__name__

        print(f"\n\n\n ============= {model_name} - {column_to_predict} ============= \n\n\n")

        # Split dataset into features (X) and target (y)
        X = features_df  # Features
        y = target_df[column_to_predict]  # Target

        # Split data into training and testing sets
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=42)

        # Initialize KFold cross-validation
        kf = KFold(n_splits=5, shuffle=True, random_state=42)

        # Find the best parameters for the model
        grid_search = GridSearchCV(estimator=model, param_grid=param_grid, cv=kf, n_jobs=-1)
        grid_search.fit(X_train, y_train)
        best_model = grid_search.best_estimator_

        # Perform cross-validation on the training set
        train_rmses, train_maes, train_mses, train_r2s = [], [], [], []
        test_rmses, test_maes, test_mses, test_r2s = [], [], [], []

        for train_index, val_index in kf.split(X_train):
            X_train_fold, X_val_fold = X_train.iloc[train_index], X_train.iloc[val_index]
            y_train_fold, y_val_fold = y_train.iloc[train_index], y_train.iloc[val_index]

            # Fit the model on the training fold
            best_model.fit(X_train_fold, y_train_fold)

            # Predict on the training fold
            y_train_pred = best_model.predict(X_train_fold)

            # Evaluate training fold
            train_rmse = root_mean_squared_error(y_train_fold, y_train_pred)
            train_mae = mean_absolute_error(y_train_fold, y_train_pred)
            train_mse = mean_squared_error(y_train_fold, y_train_pred)
            train_r2 = r2_score(y_train_fold, y_train_pred)
            train_rmses.append(train_rmse)
            train_maes.append(train_mae)
            train_mses.append(train_mse)
            train_r2s.append(train_r2)

            # Predict on the validation fold (test set for cross-validation)
            y_val_pred = best_model.predict(X_val_fold)

            # Evaluate validation fold
            test_rmse = root_mean_squared_error(y_val_fold, y_val_pred)
            test_mae = mean_absolute_error(y_val_fold, y_val_pred)
            test_mse = mean_squared_error(y_val_fold, y_val_pred)
            test_r2 = r2_score(y_val_fold, y_val_pred)
            test_rmses.append(test_rmse)
            test_maes.append(test_mae)
            test_mses.append(test_mse)
            test_r2s.append(test_r2)

        # Fit the best model on the entire training set
        best_model.fit(X_train, y_train)
        y_pred = best_model.predict(X_test)

        # LIME implementation
        # Get feature names
        feature_names = X.columns.tolist()
        explainer = LimeTabularExplainer(
            training_data=X_train.values,
            feature_names=feature_names,
            mode='regression'
        )

        idx = random.randint(1, len(X_test) - 1)
        # Explain a prediction (example with the first instance in X_test)
        exp = explainer.explain_instance(
            data_row=X_test.iloc[idx].values,
            predict_fn=lambda x: best_model.predict(pd.DataFrame(x, columns=feature_names))

        )

        # Print the explanation in the console
        fig = exp.as_pyplot_figure()
        fig.set_size_inches(17, 10)  # Adjust width and height as needed
        fig.suptitle(f'{model_name} on {column_to_predict}', fontsize=16)
        fig.savefig(utils.get_output_directory_path("plots") + f'/lime_{model_name}_{column_to_predict}_regressor.png',
                    bbox_inches='tight')

        # Feature importance and permutation
        feature_names_with_importance_coef = ['EMPTY']
        feature_names_with_importance_feature_imp = ['EMPTY']
        feature_names_with_importance_perm_imp = ['EMPTY']

        try:
            if hasattr(best_model, 'coef_'):
                importance_features = best_model.coef_
                feature_names_with_importance_coef = list(zip(X.columns, importance_features.flatten()))
                # Sort by absolute value of the importance (optional)
                feature_names_with_importance_coef.sort(key=lambda x: abs(x[1]), reverse=True)
        except Exception as e:
            print(f"Error with coef_: {e}")

        try:
            if hasattr(best_model, 'feature_importances_'):
                importance_features = best_model.feature_importances_
                feature_names_with_importance_feature_imp = list(zip(X.columns, importance_features))
                # Sort by importance (optional)
                feature_names_with_importance_feature_imp.sort(key=lambda x: x[1], reverse=True)
        except Exception as e:
            print(f"Error with feature_importances_: {e}")

        X_processed = self.preprocess_features(X.copy())
        y_processed = np.array(self.preprocess_target(y.copy()))

        try:
            best_model.fit(X_processed, y_processed)
            pi = permutation_importance(best_model, X_processed, y_processed,
                                        scoring='neg_mean_squared_error').importances_mean
            feature_names_with_importance_perm_imp = list(zip(X.columns, pi))
            # Sort by importance
            feature_names_with_importance_perm_imp.sort(key=lambda x: x[1], reverse=True)
        except Exception as e:
            print(f"Error with permutation_importance: {e}")

        # Evaluate on the test set (final evaluation)
        final_test_rmse = root_mean_squared_error(y_test, y_pred)
        final_test_mae = mean_absolute_error(y_test, y_pred)
        final_test_mse = mean_squared_error(y_test, y_pred)
        final_test_r2 = r2_score(y_test, y_pred)

        # Baseline evaluation
        baseline_predictions = [np.mean(y_train)] * len(y_test)
        baseline_rmse = root_mean_squared_error(y_test, baseline_predictions)
        baseline_mse = mean_squared_error(y_test, baseline_predictions)
        baseline_mae = mean_absolute_error(y_test, baseline_predictions)
        baseline_r2 = r2_score(y_test, baseline_predictions)

        # Generating dataframe from results
        results_df = pd.DataFrame({
            'Feature': [column_to_predict],
            'Training RMSE': [np.mean(train_rmses)],
            'Training MAE': [np.mean(train_maes)],
            'Training MSE': [np.mean(train_mses)],
            'Training R^2': [np.mean(train_r2s)],
            'Column mean': [np.mean(y_train)],

            'Test RMSE': [final_test_rmse],
            'Baseline RMSE': [baseline_rmse],
            'RMSE dif': [baseline_rmse - final_test_rmse],

            'Test MAE': [final_test_mae],
            'Baseline MAE': [baseline_mae],
            'MAE dif': [baseline_mae - final_test_mae],

            'Test MSE': [final_test_mse],
            'Baseline MSE': [baseline_mse],
            'MSE dif': [baseline_mse - final_test_mse],

            'Test R^2': [final_test_r2],
            'Baseline R^2': [baseline_r2],
            'R^2 dif': [baseline_r2 - final_test_r2],

            'Feature Importance Coef.': [feature_names_with_importance_coef],
            'Feature imp': [feature_names_with_importance_feature_imp],
            'Permutation imp': [feature_names_with_importance_perm_imp],
        })

        # Print to excel
        # Define the path to the existing Excel file
        excel_filename = utils.get_output_directory_path() + '/audio_evaluation_regressor.xlsx'

        self.append_df_to_excel(excel_filename, results_df, model_name)

    def evaluation_classifier(self, features_df, target_df, column_to_predict, model, param_grid, test_size=0.2):
        model_name = type(model).__name__
        print(f"\n\n\n ============= {model_name} - {column_to_predict} ============= \n\n\n")

        X = features_df  # Features
        y = target_df[column_to_predict]  # Target

        # Split data into training and testing sets
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=42)

        # Initialize KFold cross-validation (10 folds)
        kf = KFold(n_splits=5, shuffle=True, random_state=42)

        # Find the best parameters for model
        grid_search = GridSearchCV(estimator=model, param_grid=param_grid, cv=kf, n_jobs=-1)
        try:
            grid_search.fit(X, y)
        except Exception as e:
            print(f" Model {type(model).__name__} failed to fit")
            print(e)
            # traceback.print_exc()
            # exit()

        best_model = grid_search.best_estimator_

        # Perform cross-validation on the training set
        train_scores = []
        test_scores = []

        for train_index, test_index in kf.split(X_train):
            X_train_fold, X_val_fold = X_train.iloc[train_index], X_train.iloc[test_index]
            y_train_fold, y_val_fold = y_train.iloc[train_index], y_train.iloc[test_index]

            # Fit the model on the training fold
            best_model.fit(X_train_fold, y_train_fold)

            # Predict and evaluate on the training fold
            y_train_pred = best_model.predict(X_train_fold)
            train_score = accuracy_score(y_train_fold, y_train_pred)
            train_scores.append(train_score)

            # Predict and evaluate on the validation fold
            y_val_pred = best_model.predict(X_val_fold)
            test_score = accuracy_score(y_val_fold, y_val_pred)
            test_scores.append(test_score)

        # Evaluate on the test set (final evaluation)
        best_model.fit(X_train, y_train)
        y_pred = best_model.predict(X_test)

        # LIME implementation
        labels = ['LOW', 'MID', 'HIGH']
        class_names = y.unique()
        feature_names = X.columns.tolist()
        explainer = LimeTabularExplainer(
            training_data=X_train.values,
            class_names=labels,
            feature_names=feature_names,
            mode='classification'
        )
        idx = random.randint(1, len(X_test) - 1)
        try:
            exp = explainer.explain_instance(
                data_row=X_test.iloc[idx].values,
                predict_fn=lambda x: best_model.predict_proba(pd.DataFrame(x, columns=feature_names)),
                top_labels=len(class_names)

            ) # Print the explanation in the console
            fig = exp.as_pyplot_figure()
            fig.set_size_inches(17, 10)
            fig.suptitle(f'{model_name} on {column_to_predict}', fontsize=16)
            fig.savefig(utils.get_output_directory_path("plots") + f'/lime_{model_name}_{column_to_predict}_classifier.png',
                        bbox_inches='tight')
        except Exception as e:
            print(f"{model_name} -> Exception occurred", e)




        # Feature importance and permutation
        feature_names_with_importance_coef = ['EMPTY']
        feature_names_with_importance_feature_imp = ['EMPTY']
        feature_names_with_importance_perm_imp = ['EMPTY']

        try:
            if hasattr(best_model, 'coef_'):
                importance_features = best_model.coef_
                feature_names_with_importance_coef = list(zip(X.columns, importance_features.flatten()))
                feature_names_with_importance_coef.sort(key=lambda x: abs(x[1]), reverse=True)
        except Exception as e:
            print(f"Error with coef_: {e}")

        try:
            if hasattr(best_model, 'feature_importances_'):
                importance_features = best_model.feature_importances_
                feature_names_with_importance_feature_imp = list(zip(X.columns, importance_features))
                feature_names_with_importance_feature_imp.sort(key=lambda x: x[1], reverse=True)
        except Exception as e:
            print(f"Error with feature_importances_: {e}")

        X_processed = self.preprocess_features(X.copy())
        y_processed = np.array(self.preprocess_target(y.copy()))

        try:
            best_model.fit(X_processed, y_processed)
            pi = permutation_importance(best_model, X_processed, y_processed,
                                        scoring='neg_mean_squared_error').importances_mean
            feature_names_with_importance_perm_imp = list(zip(X.columns, pi))
            feature_names_with_importance_perm_imp.sort(key=lambda x: x[1], reverse=True)
        except Exception as e:
            print(f"Error with permutation_importance: {e}")

        # Evaluation
        accuracy = accuracy_score(y_test, y_pred)

        conf_matrix = confusion_matrix(y_test, y_pred, labels=labels)
        cm_table = pd.DataFrame(data=conf_matrix, index=labels, columns=labels)

        precision = precision_score(y_test, y_pred, average='weighted', zero_division=0)
        recall = recall_score(y_test, y_pred, average='weighted')
        f1 = f1_score(y_test, y_pred, average='weighted')
        mcc = matthews_corrcoef(y_test, y_pred)

        # Baseline evaluation
        most_frequent_value = y.value_counts().idxmax()
        baseline_predictions = [most_frequent_value] * len(y_test)

        baseline_accuracy = accuracy_score(y_test, baseline_predictions)
        baseline_precision = precision_score(y_test, baseline_predictions, average='weighted', zero_division=0)
        baseline_recall = recall_score(y_test, baseline_predictions, average='weighted')
        baseline_f1 = f1_score(y_test, baseline_predictions, average='weighted')
        baseline_mcc = matthews_corrcoef(y_test, baseline_predictions)

        # Calculate average training and validation scores
        avg_train_score = sum(train_scores) / len(train_scores)
        avg_test_score = sum(test_scores) / len(test_scores)

        # Prepare results DataFrame
        results_df = pd.DataFrame({
            'Feature': [column_to_predict],
            'Training Accuracy': [avg_train_score],
            'Validation Accuracy': [avg_test_score],

            'Test Accuracy': [accuracy],
            'Baseline Accuracy': [baseline_accuracy],

            'Precision': [precision],
            'Baseline Precision': [baseline_precision],

            'Recall': [recall],
            'Baseline Recall': [baseline_recall],

            'F1-Score': [f1],
            'Baseline F1-Score': [baseline_f1],

            'MCC': [mcc],
            'Baseline MCC': [baseline_mcc],

            'Confusion Matrix': [cm_table.to_string()],

            'Feature Importance Coef.': [feature_names_with_importance_coef],
            'Feature imp': [feature_names_with_importance_feature_imp],
            'Permutation imp': [feature_names_with_importance_perm_imp]

        })

        # Print results to Excel
        excel_filename = utils.get_output_directory_path() + '/audio_evaluation_classifier.xlsx'

        self.append_df_to_excel(excel_filename, results_df, model_name)

    def run(self):
        print("Processing audio!\n")

        print("Loading dataset!\n")

        original_df = pd.read_excel(utils.get_output_directory_path() + '/skladba_dataframe_cleaned.xlsx')

        columns_to_predict = [
            col for col in original_df.columns
            if col.lower().endswith('_mean') and (
                    col.lower().startswith('rhythm') or
                    col.lower().startswith('melody') or
                    col.lower().startswith('harmony'))

        ]

        target_df = original_df[columns_to_predict]

        # Models that are being used in ML algorithm
        regressors = [
            RandomForestRegressor(random_state=42),
            GradientBoostingRegressor(random_state=42),
            ExtraTreesRegressor(random_state=42),
            DecisionTreeRegressor(random_state=42),
            LinearRegression(),
            SVR(),
            SGDRegressor(random_state=42),
            Ridge(random_state=42),
        ]

        classifiers = [
            SVC(random_state=42),
            RandomForestClassifier(random_state=42),
            LogisticRegression(random_state=42),
            RidgeClassifier(random_state=42),
            SGDClassifier(random_state=42),
            DecisionTreeClassifier(random_state=42),
            KNeighborsClassifier()
        ]
        param_grids_regressor = [

            {  # RandomForestRegressor parameters
                'n_estimators': [50, 100, 200, 300],
                'max_depth': [None, 10, 20, 50, 100],
                'min_samples_split': [2, 5, 10],
                'min_samples_leaf': [1, 2, 5, 10],
                'bootstrap': [True, False]
            },
            {  # GradientBoostingRegressor parameters
                'n_estimators': [50, 100, 200, 300],
                'learning_rate': [0.01, 0.05, 0.1, 0.2],
                'max_depth': [3, 5, 10, 20],
                'min_samples_split': [2, 5, 10],
                'min_samples_leaf': [1, 2, 5],
                'subsample': [0.7, 0.8, 0.9, 1.0]
            },
            {  # ExtraTreesRegressor parameters
                'n_estimators': [50, 100, 200, 300],
                'max_depth': [None, 10, 20, 50],
                'min_samples_split': [2, 5, 10],
                'min_samples_leaf': [1, 2, 5, 10],
                'bootstrap': [True, False]
            },
            {  # DecisionTreeRegressor parameters
                'max_depth': [None, 10, 20, 50],
                'splitter': ['best', 'random'],
                'min_samples_split': [2, 5, 10],
                'min_samples_leaf': [1, 2, 5, 10]
            },
            {  # LinearRegression has no hyperparameters to tune
                'fit_intercept': [True, False],
            },
            {  # SVR parameters
                'kernel': ['linear', 'poly', 'rbf', 'sigmoid'],
                'C': [0.1, 1, 10, 100],
                'gamma': ['scale', 'auto', 0.001, 0.01, 0.1, 1],
                'degree': [2, 3, 4],
                'max_iter': [100, 1000, 5000],
                'epsilon': [0.01, 0.1, 1]
            },
            {  # SGDRegressor parameters
                'loss': ['squared_error', 'huber', 'epsilon_insensitive', 'squared_epsilon_insensitive'],
                'penalty': ['l2', 'l1', 'elasticnet'],
                'alpha': [0.0001, 0.001, 0.01, 0.1],
                'fit_intercept': [True, False],
                'max_iter': [1000, 5000, 10000],
                'learning_rate': ['constant', 'optimal', 'invscaling', 'adaptive'],
                'eta0': [0.001, 0.01, 0.1],
                'early_stopping': [True, False]
            },
            {  # Ridge parameters
                'alpha': [0.1, 1, 10, 100],
                'solver': ['auto', 'svd', 'cholesky', 'lsqr', 'saga']
            }
        ]

        param_grids_classifiers = [
            {  # SVC parameters
                'C': [0.1, 1],
                'kernel': ['linear', 'poly'],
                'gamma': ['scale', 'auto', 0.01, 0.1],
                'degree': [2, 7]
            },
            {  # RandomForestClassifier parameters
                'n_estimators': [10, 100, 200],
                'max_depth': [None, 10, 100, 200],
                'min_samples_split': [2, 5, 10],
                'min_samples_leaf': [1, 2, 4],
                'bootstrap': [True, False]
            },
            {  # LogisticRegression parameters
                'C': [0.01, 0.1, 1, 10, 100],
                'penalty': ['l1', 'l2', 'elasticnet', 'none'],
                'solver': ['newton-cg', 'lbfgs', 'liblinear', 'sag', 'saga'],
                'max_iter': [100, 200, 300]
            },
            {  # RidgeClassifier parameters
                'alpha': [0.1, 1.0, 10.0],
                'solver': ['auto', 'svd', 'cholesky', 'lsqr', 'sparse_cg', 'sag', 'saga']
            },
            {  # SGDClassifier parameters
                'alpha': [0.0001, 0.001, 0.01, 0.1],
                'penalty': ['l2', 'l1', 'elasticnet'],
                'max_iter': [100, 1000, 10000],
                'loss': ['hinge', 'log_loss', 'modified_huber', 'squared_hinge', 'perceptron']
            },
            {  # DecisionTreeClassifier parameters
                'max_depth': [None, 10, 100, 200],
                'min_samples_split': [2, 10],
                'min_samples_leaf': [1, 2, 10],
                'criterion': ['gini', 'entropy']
            },
            {  # KNeighborsClassifier parameters
                'n_neighbors': [3, 5, 7, 13, 17],
                'weights': ['uniform', 'distance'],
                'algorithm': ['auto', 'ball_tree', 'kd_tree', 'brute'],
                'p': [1, 2]
            }
        ]

        # Extracting features
        song_features_df = pd.read_excel(utils.get_output_directory_path() + '/song_features_output.xlsx')
        features_df = song_features_df.drop(columns=['songs'])

        print("Evaluating Audio with Regressors!\n")
        for model, param_grid in zip(regressors, param_grids_regressor):
            for column in columns_to_predict:
                self.evaluation_regressor(features_df, target_df, column, model, param_grid)

        print("Evaluating Audio with Classifiers!\n")
        classification_df = pd.read_excel(utils.get_output_directory_path() + '/audio_features_classified.xlsx')
        columns_to_predict = [col for col in classification_df.columns if col.lower().endswith('class')]
        target_df_class = classification_df[columns_to_predict]

        for model, param_grid in zip(classifiers, param_grids_classifiers):
            for column in columns_to_predict:
                self.evaluation_classifier(features_df, target_df_class, column, model, param_grid)
