import pandas as pd
from openpyxl.reader.excel import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from nltk.tokenize import word_tokenize
from sklearn.model_selection import train_test_split, KFold, GridSearchCV
from sklearn.linear_model import LinearRegression, LogisticRegression, SGDClassifier, RidgeClassifier, SGDRegressor
from sklearn.tree import DecisionTreeRegressor, DecisionTreeClassifier, DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor, ExtraTreesRegressor
from sklearn.metrics import root_mean_squared_error, r2_score, accuracy_score
from nltk.stem import WordNetLemmatizer
from nltk import PorterStemmer
from sklearn import decomposition


# HAD TO DOWNLOAD THIS FOR THE FIRST TIME
# nltk.download('punkt')
# nltk.download('stopwords')
# nltk.download('wordnet')
class TextProcessing:
    def __init__(self, use_stemming=True, n_components=20):
        print('-- Creating TextProcessing --')
        self.use_stemming = use_stemming
        self.n_components = n_components
        self.stemmer = PorterStemmer()
        self.lemmatizer = WordNetLemmatizer()
        self.pca = decomposition.PCA(n_components=self.n_components)

    def preprocess_lyrics(self, lyrics):
        tokens = word_tokenize(lyrics)
        if self.use_stemming:
            processed_tokens = [self.stemmer.stem(token) for token in tokens]
        else:
            processed_tokens = [self.lemmatizer.lemmatize(token) for token in tokens]
        return ' '.join(processed_tokens)

    def calculate_tfidf(self, df):
        # Preprocess text
        df['processed_lyrics'] = df['lyrics'].apply(self.preprocess_lyrics)

        # Initialize the TF-IDF Vectorizer
        tfidf_vectorizer = TfidfVectorizer()

        # Fit and transform the lyrics
        tfidf_matrix = tfidf_vectorizer.fit_transform(df['processed_lyrics'])

        # Create a DataFrame from the TF-IDF matrix
        tfidf_df = pd.DataFrame(tfidf_matrix.toarray(), columns=tfidf_vectorizer.get_feature_names_out())

        # Combine the original DataFrame with the TF-IDF DataFrame
        df_combined = pd.concat([df, tfidf_df], axis=1)

        # Save the combined DataFrame to an Excel file
        df_combined.to_excel('output_tables/tf_idf_dataframe.xlsx', index=False)
        return df_combined

    def dimension_reduction(self, df):
        pca_data = self.pca.fit_transform(df)
        return pca_data, self.pca

    @staticmethod
    def merge_dataframes_by_song_name(pca_df, original_df):
        merged_df = pd.merge(original_df, pca_df, on="songs", how="inner")
        return merged_df

    def append_df_to_excel(self, filename, df, sheet_name):
        try:
            # Load the existing workbook
            book = load_workbook(filename)
            if sheet_name in book.sheetnames:
                # Load existing sheet into a DataFrame
                existing_df = pd.read_excel(filename, sheet_name=sheet_name)
                # Append the new data without column names
                combined_df = pd.concat([existing_df, df], ignore_index=True)
            else:
                # If the sheet does not exist, just use the new DataFrame
                combined_df = df

            # Write to Excel
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
                if sheet_name in writer.book.sheetnames:
                    writer.book.remove(writer.book[sheet_name])
                combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
        except FileNotFoundError:
            # If the file doesn't exist, create a new one and write the DataFrame
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    def evaluation(self, df, column_to_predict, model, test_size=0.2):

        # Selecting relevant columns
        columns_to_select = [column_to_predict] + [f'PC{i}' for i in range(1, 21)]
        selected_df = df[columns_to_select]

        # Split dataset into features (X) and target (y)
        X = selected_df.drop(columns=[column_to_predict])  # Features
        y = selected_df[column_to_predict]  # Target

        # Split data into training and testing sets
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=42)

        # Initialize your model (example: Linear Regression)
        # model = LinearRegression()

        # Initialize KFold cross-validation (10 folds)
        kf = KFold(n_splits=10, shuffle=True, random_state=42)

        # Perform cross-validation on the training set
        train_mses = []
        train_r2s = []
        test_mses = []
        test_r2s = []

        for train_index, test_index in kf.split(X_train):
            X_train_fold, X_val_fold = X_train.iloc[train_index], X_train.iloc[test_index]
            y_train_fold, y_val_fold = y_train.iloc[train_index], y_train.iloc[test_index]

            # Fit the model on the training fold
            model.fit(X_train_fold, y_train_fold)

            # Predict on the training fold
            y_train_pred = model.predict(X_train_fold)

            # Evaluate training fold
            train_rmse = root_mean_squared_error(y_train_fold, y_train_pred)
            train_r2 = r2_score(y_train_fold, y_train_pred)
            train_mses.append(train_rmse)
            train_r2s.append(train_r2)

            # Predict on the validation fold (test set for cross-validation)
            y_val_pred = model.predict(X_val_fold)

            # Evaluate validation fold
            test_mse = root_mean_squared_error(y_val_fold, y_val_pred)
            test_r2 = r2_score(y_val_fold, y_val_pred)
            test_mses.append(test_mse)
            test_r2s.append(test_r2)

        # Evaluate on the test set (final evaluation)
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)
        final_test_rmse = root_mean_squared_error(y_test, y_pred)
        final_test_r2 = r2_score(y_test, y_pred)

        # Generating dataframe from results
        # Create DataFrame
        results_df = pd.DataFrame({
            'Feature': column_to_predict,
            'Training RMSE': [sum(train_mses) / len(train_mses)],
            'Training R^2': [sum(train_r2s) / len(train_r2s)],
            'Test RMSE': final_test_rmse,
            'Test R^2': final_test_r2
        })

        # Print to excel
        model_name = type(model).__name__

        # Define the path to the existing Excel file
        excel_filename = 'output_tables/model_evaluation.xlsx'

        self.append_df_to_excel(excel_filename, results_df, model_name)

        # Print results
        # print(column_to_predict, " & ", model_name, '\n --------------------------- \n')
        # print("Training Set:")
        # print(f"Root Mean Squared Error (RMSE): {sum(train_mses) / len(train_mses):.10f}")
        # print(f"R^2 Score: {sum(train_r2s) / len(train_r2s):.10f}")
        # print("\nCross-Validation (Validation Set):")
        # print(f"Root Mean Cross-Validation RMSE: {sum(test_mses) / len(test_mses):.10f}")
        # print(f"Mean Cross-Validation R^2 Score: {sum(test_r2s) / len(test_r2s):.10f}")
        # print("\nTest Set (Final Evaluation):")
        # print(f"Final Test RMSE: {final_test_rmse:.10f}")
        # print(f"Final Test R^2 Score: {final_test_r2:.10f}")
        # print("\n\n")

    def find_optimal_hyperparameters(self, df, column_to_predict, model, param_grid):
        columns_to_select = [column_to_predict] + [f'PC{i}' for i in range(1, 21)]
        selected_df = df[columns_to_select]

        # Split dataset into features (X) and target (y)
        X = selected_df.drop(columns=[column_to_predict])  # Features
        y = selected_df[column_to_predict]  # Target

        grid_search = GridSearchCV(estimator=model, param_grid=param_grid, cv=5, scoring='neg_root_mean_squared_error')
        grid_search.fit(X, y)

        # Print the best parameters and best score
        best_params = (f"Best parameters for {type(model).__name__} when predicting {column_to_predict}:\n" +
                       str(grid_search.best_params_) + '\n')
        best_result = f"Best neg_root_mean_squared_error score: {grid_search.best_score_}\n"

        # Check if file exists
        filename = 'output_tables/hyper_parameters_for_regressors.txt'

        with open(filename, 'a') as file:
            file.write(best_params + best_result + "\n\n")

        print(f"Best parameters for {type(model).__name__} when predicting {column_to_predict}:")
        print(grid_search.best_params_)
        print(f"Best neg_root_mean_squared_error score: {grid_search.best_score_}\n")

    def evaluation_classifier(self, df, column_to_predict, model, test_size=0.2):

        # Selecting relevant columns
        columns_to_select = [column_to_predict] + [f'PC{i}' for i in range(1, 21)]
        selected_df = df[columns_to_select]

        # Split dataset into features (X) and target (y)
        X = selected_df.drop(columns=[column_to_predict])  # Features
        y = selected_df[column_to_predict]  # Target

        # Split data into training and testing sets
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=42)

        # Initialize KFold cross-validation (10 folds)
        kf = KFold(n_splits=10, shuffle=True, random_state=42)

        # Perform cross-validation on the training set
        train_scores = []
        test_scores = []

        for train_index, test_index in kf.split(X_train):
            X_train_fold, X_val_fold = X_train.iloc[train_index], X_train.iloc[test_index]
            y_train_fold, y_val_fold = y_train.iloc[train_index], y_train.iloc[test_index]

            # Fit the model on the training fold
            model.fit(X_train_fold, y_train_fold)

            # Predict and evaluate on the training fold
            y_train_pred = model.predict(X_train_fold)
            train_score = accuracy_score(y_train_fold, y_train_pred)
            train_scores.append(train_score)

            # Predict and evaluate on the validation fold
            y_val_pred = model.predict(X_val_fold)
            test_score = accuracy_score(y_val_fold, y_val_pred)
            test_scores.append(test_score)

        # Evaluate on the test set (final evaluation)
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)
        final_test_score = accuracy_score(y_test, y_pred)

        # Calculate average training and validation scores
        avg_train_score = sum(train_scores) / len(train_scores)
        avg_test_score = sum(test_scores) / len(test_scores)

        # Prepare results DataFrame
        results_df = pd.DataFrame({
            'Feature': column_to_predict,
            'Training Accuracy': [avg_train_score],
            'Validation Accuracy': [avg_test_score],
            'Test Accuracy': [final_test_score]
        })

        # Print results to Excel
        excel_filename = 'output_tables/model_evaluation_classifier.xlsx'
        model_name = type(model).__name__

        self.append_df_to_excel(excel_filename, results_df, model_name)

    @staticmethod
    def merge_dataframes(pca_df, original_df):
        merged_df = pd.merge(original_df, pca_df, on="songs", how="inner")
        merged_df.to_excel('output_tables/merged_dataframe.xlsx', index=False)
        return merged_df

    @staticmethod
    def extract_song_name(song):
        if ':' in song:
            return song.split(':')[1].strip()
        else:
            return song  # return as is if no ':' found

    def run(self):
        # Read input DataFrame
        df = pd.read_excel("output_tables/lyrics_output.xlsx")

        # Calculate TF-IDF
        full_df = self.calculate_tfidf(df)

        # Perform dimension reduction
        send_df = full_df.iloc[:, 4:]  # Selecting all columns from index 4 to the end
        pca_data, pca = self.dimension_reduction(send_df)

        # Create a DataFrame with the PCA data
        pca_df = pd.DataFrame(pca_data, columns=[f'PC{i + 1}' for i in range(pca_data.shape[1])])

        # Include the original song titles in the PCA DataFrame
        pca_df = pd.concat([full_df[['songs']], pca_df], axis=1)
        pca_df.to_excel('output_tables/PCA_dataframe.xlsx', index=False)

        # Prediction

        pca_df = pd.read_excel("output_tables/PCA_dataframe.xlsx")

        original_df = pd.read_excel("output_tables/skladba_dataframe_cleaned.xlsx",
                                    converters={'songs': self.extract_song_name})
        lyrics_columns = [col for col in original_df.columns if 'lyrics' in col.lower() or 'songs' in col.lower()]

        original_df_parameter = original_df[lyrics_columns]
        merged = self.merge_dataframes(pca_df, original_df_parameter)

        columns_to_predict = ['lyrics_complexity_mean', 'lyrics_comprehensibility_mean', 'lyrics_complexity_mean']

        # Models that are being used in ML algorithm
        models = [
            RandomForestRegressor(n_estimators=100, max_depth=None, min_samples_split=2, min_samples_leaf=1,
                                  random_state=42),
            GradientBoostingRegressor(n_estimators=100, learning_rate=0.1, max_depth=3, min_samples_split=2,
                                      min_samples_leaf=1, random_state=42),
            ExtraTreesRegressor(n_estimators=100, max_depth=None, min_samples_split=2, min_samples_leaf=1,
                                random_state=42),
            DecisionTreeRegressor(max_depth=None, min_samples_split=2, min_samples_leaf=1, random_state=42),
            LinearRegression()

        ]

        classifiers = [
            LogisticRegression(C=1.0, penalty='l2', random_state=42),
            SGDRegressor(alpha=0.0001, penalty='l2', loss='squared_loss', random_state=42),
            RidgeClassifier(alpha=1.0, random_state=42),
            SGDClassifier(alpha=0.0001, penalty='l2', loss='log', random_state=42),
            DecisionTreeClassifier(max_depth=None, min_samples_split=2, min_samples_leaf=1, random_state=42)]

        param_grids = [
            {  # RandomForestRegressor parameters
                'n_estimators': [50, 100, 200],
                'max_depth': [None, 10, 20],
                'min_samples_split': [2, 5],
                'min_samples_leaf': [1, 2]
            },
            {  # GradientBoostingRegressor parameters
                'n_estimators': [50, 100, 200],
                'learning_rate': [0.05, 0.1, 0.2],
                'max_depth': [3, 5],
                'min_samples_split': [2, 5],
                'min_samples_leaf': [1, 2]
            },
            {  # ExtraTreesRegressor parameters
                'n_estimators': [50, 100, 200],
                'max_depth': [None, 10, 20],
                'min_samples_split': [2, 5],
                'min_samples_leaf': [1, 2]
            },
            {  # DecisionTreeRegressor parameters
                'max_depth': [None, 10, 20],
                'min_samples_split': [2, 5],
                'min_samples_leaf': [1, 2]
            },
            {  # LinearRegression has no hyperparameters to tune
                'fit_intercept': [True, False]
            }
        ]

        for model, param_grid in zip(models, param_grids):
            for column in columns_to_predict:
                self.find_optimal_hyperparameters(merged, column, model, param_grid)

        for model in models:
            for column in columns_to_predict:
                self.evaluation(merged, column, model)

        # for model in models:
        #     for column in columns_to_predict:
        #         self.evaluation_classifier(merged, column, model)
